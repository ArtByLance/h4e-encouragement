#!/usr/bin/env python3
"""Create a prototype multilingual workbook from the current English review sheet."""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "eventKey",
    "label",
    "triggerDescription",
    "active",
    "templateKey",
    "message",
    "status",
    "notes",
]
LOCALES = ("en", "es", "ar")
RECIPIENT_PLACEHOLDER = "{{recipientName}}"
LEGACY_PLACEHOLDER = "[Name]"

TRANSLATIONS = {
    ("completed_day_1", "template1"): {
        "es": "Bienvenido al camino, {{recipientName}}! Vas muy bien. Nos vemos el proximo dia!",
        "ar": "اهلا بك في الرحلة يا {{recipientName}}! انت تتقدم بشكل رائع. نراك في اليوم التالي!",
    },
    ("completed_day_1", "template2"): {
        "es": "Buen comienzo, {{recipientName}}! Completaste el primer dia y ya estas tomando impulso. Nos vemos el proximo dia!",
        "ar": "بداية موفقة يا {{recipientName}}! لقد اكملت اليوم الاول وبدأت تبني زخما حقيقيا. نراك في اليوم التالي!",
    },
    ("completed_day_1", "template3"): {
        "es": "{{recipientName}}, ya terminaste el primer dia! Es un gran primer paso. Sigue presente y nos vemos el proximo dia!",
        "ar": "{{recipientName}}، لقد انتهى اليوم الاول! هذه خطوة اولى قوية. استمر في الحضور ونراك في اليوم التالي!",
    },
    ("completed_3_days", "template1"): {
        "es": "{{recipientName}}, lo estas haciendo excelente! Tres dias seguidos y estas construyendo algo solido. Sigue adelante hoy!",
        "ar": "{{recipientName}}، اداؤك ممتاز! ثلاثة ايام وانت تبني شيئا قويا. واصل التقدم اليوم!",
    },
    ("completed_3_days", "template2"): {
        "es": "Muy buen trabajo, {{recipientName}}! Tres dias completados y estas logrando progreso real. Mantente constante hoy!",
        "ar": "عمل رائع يا {{recipientName}}! اكملت ثلاثة ايام وتحقق تقدما حقيقيا. حافظ على الاستمرارية اليوم!",
    },
    ("completed_3_days", "template3"): {
        "es": "{{recipientName}}, te has presentado durante tres dias y eso importa. Mantengamos la racha hoy!",
        "ar": "{{recipientName}}، لقد التزمت لثلاثة ايام وهذا مهم. لنحافظ على هذا التقدم اليوم!",
    },
    ("reached_50_percent", "template1"): {
        "es": "{{recipientName}}, llegaste a la mitad! Eso es increible. Vas muy bien, sigamos avanzando hoy!",
        "ar": "{{recipientName}}، وصلت الى منتصف الطريق! هذا انجاز رائع. انت تتقدم بشكل ممتاز، فلنواصل اليوم!",
    },
    ("reached_50_percent", "template2"): {
        "es": "Ya vas a la mitad, {{recipientName}}! Has avanzado mucho. Mantengamos ese progreso hoy!",
        "ar": "انت في منتصف الطريق يا {{recipientName}}! لقد انجزت الكثير. لنحافظ على هذا التقدم اليوم!",
    },
    ("reached_50_percent", "template3"): {
        "es": "{{recipientName}}, estas en la mitad del camino y eso es muy importante. Sigamos avanzando desde aqui!",
        "ar": "{{recipientName}}، وصلت الى نقطة المنتصف وهذا امر مهم جدا. لنكمل التقدم من هنا!",
    },
    ("reached_75_percent", "template1"): {
        "es": "{{recipientName}}, estas avanzando con mucha fuerza! Ya llegaste muy lejos y vas muy bien, terminemos con energia!",
        "ar": "{{recipientName}}، انت تحقق تقدما قويا جدا! وصلت بعيدا واداؤك رائع، فلننهي بقوة!",
    },
    ("reached_75_percent", "template2"): {
        "es": "Ya completaste el 75%, {{recipientName}}! La meta esta cerca. Sigue avanzando con fuerza!",
        "ar": "اكملت 75% يا {{recipientName}}! خط النهاية اصبح قريبا. استمر بقوة!",
    },
    ("reached_75_percent", "template3"): {
        "es": "{{recipientName}}, has llegado muy lejos y estas haciendo un gran trabajo. Sigamos hasta terminar con fuerza!",
        "ar": "{{recipientName}}، وصلت بعيدا وتقوم بعمل ممتاز. لنكمل الطريق وننهي بقوة!",
    },
    ("completed_course", "template1"): {
        "es": "{{recipientName}}, lo lograste! Es un gran paso! Te mantuviste constante y terminaste con fuerza. Debes sentir orgullo!",
        "ar": "{{recipientName}}، لقد فعلتها! هذا انجاز كبير! استمريت حتى النهاية وانهيت بقوة. من حقك ان تفخر!",
    },
    ("completed_course", "template2"): {
        "es": "Felicidades, {{recipientName}}! Completaste el curso y cumpliste hasta el final. Es algo para celebrar!",
        "ar": "تهانينا يا {{recipientName}}! اكملت الدورة وواصلت حتى النهاية. هذا انجاز يستحق الاحتفال!",
    },
    ("completed_course", "template3"): {
        "es": "{{recipientName}}, terminaste el curso! Tu constancia dio resultado. Toma un momento para celebrar este logro!",
        "ar": "{{recipientName}}، لقد انهيت الدورة! استمراريتك اثمرت. خذ لحظة للاحتفال بهذا الانجاز!",
    },
    ("completed_every_5_days", "template1"): {
        "es": "{{recipientName}}, vas muy bien! Sigues presentandote y eso esta funcionando. Sigamos adelante hoy!",
        "ar": "{{recipientName}}، انت تسير بشكل رائع! تستمر في الحضور وهذا يصنع فرقا. لنواصل اليوم!",
    },
    ("completed_every_5_days", "template2"): {
        "es": "Muy buena constancia, {{recipientName}}! Cada vez que te presentas, todo suma. Mantengamos el progreso hoy!",
        "ar": "استمرارية رائعة يا {{recipientName}}! كل مرة تلتزم فيها تضيف الى تقدمك. لنحافظ على التقدم اليوم!",
    },
    ("completed_every_5_days", "template3"): {
        "es": "{{recipientName}}, tu esfuerzo constante esta marcando la diferencia. Mantente firme y sigue adelante hoy!",
        "ar": "{{recipientName}}، مجهودك المستمر يصنع فرقا. اثبت وواصل التقدم اليوم!",
    },
    ("inactive_3_days", "template1"): {
        "es": "{{recipientName}}, han pasado unos dias. Es buen momento para retomar y volver a avanzar. Tu puedes!",
        "ar": "{{recipientName}}، مرت بضعة ايام. هذا وقت مناسب للعودة والتحرك من جديد. انت قادر!",
    },
    ("inactive_3_days", "template2"): {
        "es": "{{recipientName}}, pasaron unos dias, pero puedes recomenzar ahora. Retoma con un paso hoy!",
        "ar": "{{recipientName}}، مرت بضعة ايام، لكن يمكنك البدء من جديد الان. ارجع بخطوة واحدة اليوم!",
    },
    ("inactive_3_days", "template3"): {
        "es": "Volvamos a avanzar, {{recipientName}}. Solo hace falta un dia para empezar a recuperar impulso!",
        "ar": "لنعد للتحرك يا {{recipientName}}. يوم واحد يكفي لتبدأ في استعادة الزخم!",
    },
    ("inactive_5_days", "template1"): {
        "es": "{{recipientName}}, ya paso un poco de tiempo. Retomemos. Haz solo un dia y volveras al camino!",
        "ar": "{{recipientName}}، مر بعض الوقت. لنعد من جديد. اكمل يوما واحدا وستعود الى المسار!",
    },
    ("inactive_5_days", "template2"): {
        "es": "{{recipientName}}, todavia puedes retomarlo. Empieza con un dia y vuelve a moverte!",
        "ar": "{{recipientName}}، ما زال بامكانك العودة. ابدأ بيوم واحد وتحرك من جديد!",
    },
    ("inactive_5_days", "template3"): {
        "es": "Es momento de un nuevo comienzo, {{recipientName}}. Completa un dia hoy y volveras a avanzar!",
        "ar": "حان وقت بداية جديدة يا {{recipientName}}. اكمل يوما واحدا اليوم وستعود للتقدم!",
    },
    ("inactive_10_days", "template1"): {
        "es": "{{recipientName}}, es momento de volver a empezar! Retoma con solo un dia. Tu puedes hacerlo!",
        "ar": "{{recipientName}}، حان وقت العودة من جديد! ابدأ بيوم واحد فقط. انت تستطيع!",
    },
    ("inactive_10_days", "template2"): {
        "es": "{{recipientName}}, no te preocupes por la pausa. Vuelve por un dia y reinicia tu progreso!",
        "ar": "{{recipientName}}، لا تقلق من فترة التوقف. ارجع ليوم واحد واعد تشغيل تقدمك!",
    },
    ("inactive_10_days", "template3"): {
        "es": "Puedes empezar de nuevo hoy, {{recipientName}}. Un dia es suficiente para volver al camino!",
        "ar": "يمكنك البدء من جديد اليوم يا {{recipientName}}. يوم واحد يكفي للعودة الى المسار!",
    },
}


def clean(value) -> str:
    return str(value if value is not None else "").strip()


def convert_placeholder(message: str) -> str:
    return clean(message).replace(LEGACY_PLACEHOLDER, RECIPIENT_PLACEHOLDER)


def read_source_rows(path: Path):
    workbook = load_workbook(path, data_only=True)
    try:
        sheet = workbook.active
        headers = {clean(cell.value): index for index, cell in enumerate(sheet[1])}
        template_columns = sorted(
            [header for header in headers if header.startswith("template")],
            key=lambda value: int(value.replace("template", "")),
        )

        rows = []
        for source_row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(clean(value) for value in source_row):
                continue

            for template_key in template_columns:
                event_key = clean(source_row[headers["eventKey"]])
                rows.append(
                    {
                        "eventKey": event_key,
                        "label": clean(source_row[headers["label"]]),
                        "triggerDescription": clean(source_row[headers["triggerDescription"]]),
                        "active": source_row[headers["active"]],
                        "templateKey": template_key,
                        "message": convert_placeholder(source_row[headers[template_key]]),
                    }
                )

        return rows
    finally:
        workbook.close()


def style_sheet(sheet, locale: str):
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:H{sheet.max_row}"
    sheet.sheet_view.rightToLeft = locale == "ar"

    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="top", wrap_text=True)

    widths = {
        "A": 24,
        "B": 24,
        "C": 34,
        "D": 10,
        "E": 14,
        "F": 78,
        "G": 14,
        "H": 36,
    }

    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    validation = DataValidation(type="list", formula1='"approved,draft,review"', allow_blank=False)
    sheet.add_data_validation(validation)
    validation.add(f"G2:G{sheet.max_row}")


def create_workbook(rows, output_path: Path):
    workbook = Workbook()
    workbook.remove(workbook.active)

    instructions = workbook.create_sheet("instructions")
    instructions.append(["Prototype multilingual workbook"])
    instructions.append(["Use en/es/ar tabs for import."])
    instructions.append(["Preserve {{recipientName}} exactly in translated messages."])
    instructions.append(["English rows are approved; Spanish and Arabic rows are draft for review."])

    for locale in LOCALES:
        sheet = workbook.create_sheet(locale)
        sheet.append(HEADERS)

        for row in rows:
            event_key = row["eventKey"]
            template_key = row["templateKey"]

            if locale == "en":
                message = row["message"]
                status = "approved"
                notes = "Approved English source from prototype workbook."
            else:
                message = TRANSLATIONS[(event_key, template_key)][locale]
                status = "draft"
                notes = "Prototype translation; requires client/native review."

            sheet.append(
                [
                    event_key,
                    row["label"],
                    row["triggerDescription"],
                    row["active"],
                    template_key,
                    message,
                    status,
                    notes,
                ]
            )

        style_sheet(sheet, locale)

    workbook.save(output_path)
    workbook.close()


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create a prototype multilingual workbook.")
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("msg-templates.xlsx"),
        help="Old-format English source workbook",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("msg-templates-prototype.xlsx"),
        help="Prototype multilingual workbook output path",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    rows = read_source_rows(args.source)
    create_workbook(rows, args.output)
    print(f"Wrote {len(rows)} templates to {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
