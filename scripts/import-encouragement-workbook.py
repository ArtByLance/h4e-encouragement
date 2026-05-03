#!/usr/bin/env python3
"""Import an approved multilingual encouragement workbook into JSON."""

from __future__ import annotations

import argparse
import json
import re
import sys
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_LOCALES = ("en", "es", "ar")
REQUIRED_COLUMNS = (
    "eventKey",
    "label",
    "triggerDescription",
    "active",
    "templateKey",
    "message",
    "status",
    "notes",
)
ALLOWED_PLACEHOLDERS = {"recipientName"}
PLACEHOLDER_PATTERN = re.compile(r"{{\s*([^{}\s]+)\s*}}")


def clean(value) -> str:
    return str(value if value is not None else "").strip()


def parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value

    normalized = clean(value).lower()
    return normalized in {"1", "true", "yes", "y", "active"}


def get_headers(sheet) -> dict[str, int]:
    headers = {clean(cell.value): index for index, cell in enumerate(sheet[1])}
    missing = [column for column in REQUIRED_COLUMNS if column not in headers]

    if missing:
        raise ValueError(
            f'Sheet "{sheet.title}" is missing required columns: {", ".join(missing)}'
        )

    return headers


def validate_placeholders(message: str, context: str) -> set[str]:
    placeholders = set(PLACEHOLDER_PATTERN.findall(message))
    unknown = sorted(placeholders - ALLOWED_PLACEHOLDERS)

    if unknown:
        raise ValueError(
            f'{context} contains unknown placeholders: {", ".join(unknown)}'
        )

    return placeholders


def iter_rows(sheet, headers: dict[str, int]):
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        values = {
            column: row[headers[column]] if headers[column] < len(row) else None
            for column in REQUIRED_COLUMNS
        }

        if not any(clean(value) for value in values.values()):
            continue

        yield row_index, values


def read_workbook(path: Path, include_unapproved: bool):
    workbook = load_workbook(path, data_only=True)

    try:
        return read_loaded_workbook(workbook, include_unapproved)
    finally:
        workbook.close()


def read_loaded_workbook(workbook, include_unapproved: bool):
    missing_sheets = [locale for locale in REQUIRED_LOCALES if locale not in workbook.sheetnames]

    if missing_sheets:
        raise ValueError(f"Workbook is missing required sheets: {', '.join(missing_sheets)}")

    messages: OrderedDict[str, dict] = OrderedDict()
    placeholder_sets: dict[tuple[str, str, str], set[str]] = {}
    english_placeholders: dict[tuple[str, str], set[str]] = {}
    seen: set[tuple[str, str, str]] = set()

    for locale in REQUIRED_LOCALES:
        sheet = workbook[locale]
        headers = get_headers(sheet)

        for row_index, row in iter_rows(sheet, headers):
            status = clean(row["status"]).lower()

            if not include_unapproved and status != "approved":
                continue

            event_key = clean(row["eventKey"])
            template_key = clean(row["templateKey"])
            message = clean(row["message"])
            context = f'{locale}!row {row_index} ({event_key or "missing eventKey"})'

            if not event_key:
                raise ValueError(f"{context} is missing eventKey")
            if not template_key:
                raise ValueError(f"{context} is missing templateKey")
            if not message:
                raise ValueError(f"{context} is missing message")

            row_id = (locale, event_key, template_key)
            if row_id in seen:
                raise ValueError(
                    f'{context} duplicates template "{template_key}" for event "{event_key}"'
                )
            seen.add(row_id)

            placeholders = validate_placeholders(message, context)
            placeholder_sets[row_id] = placeholders

            if locale == "en":
                english_placeholders[(event_key, template_key)] = placeholders

            if event_key not in messages:
                messages[event_key] = {
                    "label": clean(row["label"]) or event_key,
                    "triggerDescription": clean(row["triggerDescription"]),
                    "active": parse_bool(row["active"]),
                    "templates": {locale_key: OrderedDict() for locale_key in REQUIRED_LOCALES},
                }

            messages[event_key]["templates"][locale][template_key] = message

    for event_key, definition in messages.items():
        if definition["active"] and not definition["templates"]["en"]:
            raise ValueError(f'Active event "{event_key}" is missing an English fallback')

    for locale in REQUIRED_LOCALES:
        if locale == "en":
            continue

        for event_key, definition in messages.items():
            for template_key in definition["templates"][locale]:
                english_key = (event_key, template_key)
                localized_key = (locale, event_key, template_key)

                if english_key not in english_placeholders:
                    raise ValueError(
                        f'{locale} template "{event_key}/{template_key}" has no English fallback'
                    )

                if placeholder_sets[localized_key] != english_placeholders[english_key]:
                    raise ValueError(
                        f'{locale} template "{event_key}/{template_key}" does not match English placeholder usage'
                    )

    return messages


def build_catalog(messages: OrderedDict[str, dict]) -> dict:
    return {
        "schemaVersion": "2.0.0",
        "catalogId": "attendee-encouragement",
        "catalogName": "Attendee Encouragement Messages",
        "description": (
            "Runtime catalog for event-driven attendee encouragement messages. "
            "Application logic should use stable event keys; labels and trigger "
            "descriptions are for admin UI, QA, and operational review."
        ),
        "defaultLocale": "en",
        "supportedLocales": list(REQUIRED_LOCALES),
        "locales": {
            "en": {"label": "English", "direction": "ltr"},
            "es": {"label": "Spanish", "direction": "ltr"},
            "ar": {"label": "Arabic", "direction": "rtl"},
        },
        "placeholders": {"recipientName": "{{recipientName}}"},
        "defaults": {
            "recipientName": "Mike",
            "phoneNumber": "678-777-7100",
            "eventKey": "inactive_3_days" if "inactive_3_days" in messages else next(iter(messages), ""),
            "locale": "en",
        },
        "messageOrder": list(messages.keys()),
        "messages": messages,
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Import a multilingual encouragement workbook into JSON."
    )
    parser.add_argument("workbook", type=Path, help="Path to the source .xlsx workbook")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=Path("msg-data.json"),
        help="Output JSON path",
    )
    parser.add_argument(
        "--include-unapproved",
        action="store_true",
        help="Import all non-empty rows instead of only rows with status=approved",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        messages = read_workbook(args.workbook, args.include_unapproved)
        catalog = build_catalog(messages)
        args.output.write_text(
            json.dumps(catalog, ensure_ascii=False, indent=2) + "\n",
            encoding="utf-8",
        )
    except Exception as error:  # noqa: BLE001 - command line script should print concise failures.
        print(f"Import failed: {error}", file=sys.stderr)
        return 1

    print(f"Imported {len(messages)} events into {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
